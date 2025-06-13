#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_bopo2_in_excel.py

將 Excel 檔案「漢字庫」工作表中的「台羅音標」／TL 欄批次轉成「注音二式」，寫入「注音二式」欄（不存在就新增）。
用法：
    python update_bopo2_in_excel.py input.xlsx [output.xlsx]

執行完後，請打開 Excel 手動查核「注音二式」欄是否正確。
"""
import re
import sys

from openpyxl import load_workbook

# 聲母（初聲）對應表，從長到短匹配
TL_INITIAL_MAP = {
    'tshi': 'ci',   # ㄑ
    'tsi':  'zi',   # ㄐ
    'tsh':  'c',    # ㄘ
    'ts':   'z',    # ㄗ
    'ji':   'zzi',  # ㆢ
    'j':    'zz',   # ㆡ
    'b':   'bb',    # ㆠ
    'ph':   'p',    # ㄆ
    'p':    'b',    # ㄅ
    'th':   't',    # ㄊ
    't':    'd',    # ㄉ
    'kh':   'k',    # ㄎ
    'k':    'g',    # ㄍ
    'g':    'gg',   # ㆣ
    # 'ng':   'ng-',  # ㄫ（若有初聲 ng）
    'm':    'm',    # ㄇ
    'n':    'n',    # ㄋ
    'l':    'l',    # ㄌ
    'h':    'h',    # ㄏ
    's':    's',    # ㄙ
    'si':   'si',   # ㄒ （注意放在 's' 前面）
}

# 只轉這三組韻母
TL_FINAL_MAP = {
    'onn': 'oonn',  # ㆧ
    'o':   'or',    # ㄜ
    'io':  'ior',   # io→ior
}

def convert_tl_to_bopo2(code: str) -> str:
    m = re.match(r'^([a-z]+)(\d+)$', code or '')
    if not m:
        return code or ''
    body, tone = m.group(1), m.group(2)
    onset, rest = '', body
    for key in sorted(TL_INITIAL_MAP, key=lambda x: -len(x)):
        if body.startswith(key):
            onset = TL_INITIAL_MAP[key]
            rest  = body[len(key):]
            break
    # 只轉三組
    if rest in TL_FINAL_MAP:
        rest = TL_FINAL_MAP[rest]
    return f"{onset}{rest}{tone}"

def main(inp, outp=None):
    wb = load_workbook(inp)
    if '漢字庫' not in wb.sheetnames:
        print("找不到工作表「漢字庫」"); sys.exit(1)
    ws = wb['漢字庫']

    # 找標題行
    hdr = next(ws.iter_rows(min_row=1, max_row=1))
    cols = {c.value: c.col_idx for c in hdr}
    if '台羅音標' not in cols:
        print("找不到欄「台羅音標」"); sys.exit(1)
    src = cols['台羅音標']
    if '注音二式' in cols:
        dst = cols['注音二式']
    else:
        dst = ws.max_column + 1
        ws.cell(row=1, column=dst, value='注音二式')

    # 逐列轉
    for r in range(2, ws.max_row+1):
        raw = ws.cell(r, src).value
        if raw:
            new = convert_tl_to_bopo2(str(raw).strip())
            ws.cell(r, dst, new)

    save_path = outp or inp
    wb.save(save_path)
    print(f"已將「注音二式」欄更新到：{save_path}")

if __name__ == '__main__':
    if len(sys.argv)==1:
        print("用法：python update_bopo2_in_excel.py input.xlsx [output.xlsx]"); sys.exit(0)
    inp = sys.argv[1]
    outp= sys.argv[2] if len(sys.argv)>2 else None
    main(inp, outp)
