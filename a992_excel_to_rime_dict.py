#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_to_rime_dict.py

將 Excel「漢字庫」工作表中的下列欄位：
  - 漢字
  - 注音二式
  - 常用度
  - 摘要說明
輸出成 Rime 格式的字典檔 zu_im_2.dict.yaml。

用法：
    python excel_to_rime_dict.py input.xlsx [output.dict.yaml]
"""
import sys

from openpyxl import load_workbook


def main(inp, outp='zu_im_2.dict.yaml'):
    wb = load_workbook(inp, data_only=True)
    if '漢字庫' not in wb.sheetnames:
        print("找不到工作表「漢字庫」"); sys.exit(1)
    ws = wb['漢字庫']

    # 標題行
    hdr = next(ws.iter_rows(min_row=1, max_row=1))
    cols = {c.value: c.col_idx for c in hdr}
    for key in ('漢字','注音二式','常用度','摘要說明'):
        if key not in cols:
            print(f"找不到欄「{key}」"); sys.exit(1)

    # 開頭固定格式
    header = """# Rime dictionary
# encoding: utf-8
#
# 河洛白話音
#
---
name: zu_im_2
version: "v0.1.0.0"
sort: by_weight
use_preset_vocabulary: false
columns:
  - text    # 漢字
  - code    # 台灣注音二式
  - weight  # 常用度（優先顯示度）
  - stem    # 用法舉例
import_tables:
  # - tl_ji_khoo_kah_kut_bun
...
"""
    with open(outp, 'w', encoding='utf-8') as fout:
        fout.write(header)
        # 逐列寫入
        for r in range(2, ws.max_row+1):
            ch = ws.cell(r, cols['漢字']).value
            code = ws.cell(r, cols['注音二式']).value
            w = ws.cell(r, cols['常用度']).value
            s = ws.cell(r, cols['摘要說明']).value or ''
            if not ch or not code:
                continue
            # weight 若是小數，去掉前後空白
            w_str = str(w).strip()
            fout.write(f"{ch}\t{code}\t{w_str}\t{s}\n")
    print(f"已匯出 Rime 字典：{outp}")

if __name__ == '__main__':
    if len(sys.argv)<2:
        print("用法：python excel_to_rime_dict.py input.xlsx [output.dict.yaml]"); sys.exit(0)
    inp = sys.argv[1]
    outp= sys.argv[2] if len(sys.argv)>2 else 'zu_im_2.dict.yaml'
    main(inp, outp)
