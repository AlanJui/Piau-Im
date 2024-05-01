import argparse
import os
import os.path
import time

import xlwings as xw
from openpyxl import load_workbook


def get_cmd_input():
    parser = argparse.ArgumentParser(description='Process some files.')
    parser.add_argument('-i', '--input', default='Piau_Tsu_Im', help='Input file name')
    parser.add_argument('-o', '--output', default='', help='Output file name')
    args = parser.parse_args()

    return {
        "input": args.input,
        "output": args.output,
    }


def open_excel_file(main_file_name):
    # 檢查檔案名稱是否已包含副檔名
    file_name, file_extension = os.path.splitext(main_file_name)
    if not file_extension:
        # 如果沒有副檔名，添加 .xlsx
        excel_file_name = file_name + '.xlsx'
    else:
        excel_file_name = main_file_name

    current_path = os.getcwd()
    # file_path = os.path.join(current_path, "output", excel_file_name)
    file_path = os.path.join(current_path, excel_file_name)
    return xw.Book(file_path)


#==================================================================
# 程式碼流程：
# 1. 檢查提供的文件名是否包含副檔名，如果沒有則自動添加 .xlsx。
# 2. 打開原始 Excel 文件。
# 3. 另存為一份副本到 output 子目錄下，命名為 Piau-Tsu-Im.xlsx。
# 4. 關閉原始工作簿。
# 5. 重新打開新保存的副本。
# 6. 最後返回這個新打開的工作簿對象。
#==================================================================
def save_to_a_working_copy(main_file_name):
    global wb1
    # 检查文件名称是否已包含扩展名
    file_name, file_extension = os.path.splitext(main_file_name)
    if not file_extension:
        # 如果没有扩展名，添加 .xlsx
        excel_file_name = file_name + '.xlsx'
    else:
        excel_file_name = main_file_name

    # 获取当前工作目录并构建原始文件的完整路径
    current_path = os.getcwd()
    file_path = os.path.join(current_path, "output", excel_file_name)

    # 确保 output 文件夹存在
    if not os.path.exists(os.path.dirname(file_path)):
        os.makedirs(os.path.dirname(file_path))

    # 尝试打开 Excel 文件
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        print(f"File {file_path} not found.")
        return None

    try:
        # 指定新保存路径和新文件名
        # new_file_path = os.path.join(current_path, "output", "Piau-Tsu-Im.xlsx")
        new_file_path = os.path.join(current_path, "working.xlsx")

        # 使用另存为将文件保存至指定路径
        wb.save(new_file_path)
    finally:
        # 无论是否成功，都关闭原始工作簿
        wb.close()

    # 重新打开刚才另存的工作簿
    # new_wb = load_workbook(new_file_path)

    # 返回新的工作簿对象
    # return new_wb


def write_to_excel_file(excel_workbook):
    # 自工作表「env」取得新檔案名稱
    setting_sheet = excel_workbook.sheets["env"]
    new_file_name = str(
        setting_sheet.range("C4").value
    ).strip()
    current_path = os.getcwd()
    new_file_path = os.path.join(
        current_path,
        "output", 
        f"【河洛話注音】{new_file_name}" + ".xlsx")
    print(f"儲存輸出，置於檔案：{new_file_path}")

    # 在儲存文件前確保 output 資料夾存在。如果不存在，則先創建它
    output_dir = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 儲存新建立的工作簿
    excel_workbook.save(new_file_path)

    # 等待一段時間讓 save 完成
    time.sleep(3)


def close_excel_file(excel_workbook):
    # 關閉工作簿
    excel_workbook.close()

# -----------------------------------------------------------------
# 將「字串」轉換成「串列（Characters List）」
# Python code to convert string to list character-wise
# -----------------------------------------------------------------
def convert_string_to_chars_list(string):
    list1 = []
    list1[:0] = string
    return list1

# -----------------------------------------------------------------
# 要生成超連結的目錄
# directory = 'output'
# extenstion = 'xlsx'
# exculude_list = ['Piau-Tsu-Im.xlsx', 'env.xlsx', 'env_osX.xlsx']
# -----------------------------------------------------------------
def create_file_list(directory, extension, exculude_list):
    # 建立檔案清單
    file_list = []

    # 遍歷目錄下的檔案
    for filename in os.listdir(directory):
        # 排除 index.html 和 _template.html 檔案
        if filename not in exculude_list:
            if filename.endswith(extension):
                file_list.append(filename)

    return file_list

    

# -----------------------------------------------------
# Backup the original file
# -----------------------------------------------------
# def open_excel_file(main_file_name):
#     # 檢查檔案名稱是否已包含副檔名
#     file_name, file_extension = os.path.splitext(main_file_name)
#     if not file_extension:
#         # 如果沒有副檔名，添加 .xlsx
#         excel_file_name = file_name + '.xlsx'
#     else:
#         excel_file_name = main_file_name

#     # 獲取當前工作目錄並構建原始檔案的完整路徑
#     current_path = os.getcwd()
#     file_path = os.path.join(current_path, "output", excel_file_name)

#     # 確保 output 資料夾存在
#     if not os.path.exists(os.path.dirname(file_path)):
#         os.makedirs(os.path.dirname(file_path))

#     # 嘗試打開 Excel 文件
#     try:
#         wb = xw.Book(file_path)
#     except FileNotFoundError:
#         print(f"File {file_path} not found.")
#         return None

#     # 指定新存檔路徑和新檔名
#     # new_file_path = os.path.join(current_path, "output", "Piau-Tsu-Im.xlsx")
#     # new_file_path = os.path.join(current_path, "output", "_tmp.xlsx")
#     new_file_path = os.path.join(current_path, "working.xlsx")

#     # 使用另存為將檔案儲存至指定路徑
#     wb.save(new_file_path)

#     # 關閉原始工作簿
#     wb.close()

#     # 重新打開剛才另存的工作簿
#     new_wb = xw.Book(new_file_path)

#     # 回傳新的工作簿對象
#     return new_wb