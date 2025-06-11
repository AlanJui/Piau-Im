#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
convert_tl_to_bopo2_excel.py

將 Excel 檔案中「漢字庫」工作表的「台羅音標」轉成「注音二式」。
若原檔案沒有「注音二式」欄，程式會自動新增。
預設會覆蓋輸入檔，也可指定第二個參數另存輸出檔。

用法：
    python convert_tl_to_bopo2_excel.py input.xlsx [output.xlsx]
"""

import re
import sys

from openpyxl import load_workbook

# 全部聲母對照：TL → 注音二式
TL_INITIAL_MAP = {
    # 'tshi': 'ci',   # ㄑ
    # 'tsi':  'zi',   # ㄐ
    'tsh':  'c',    # ㄘ
    'ts':   'z',    # ㄗ
    'ji':   'zzi',  # ㆢ
    # 'j':    'zz',   # ㆡ
    'b':   'bb',    # ㆠ
    'ph':   'p',    # ㄆ
    'p':    'b',    # ㄅ
    'th':   't',    # ㄊ
    't':    'd',    # ㄉ
    'kh':   'k',    # ㄎ
    'k':    'g',    # ㄍ
    'g':    'gg',   # ㆣ
    # 'ng':   'ng-',  # ㄫ（若有初聲 ng）
    'm':    'm',    # ㄇ
    'n':    'n',    # ㄋ
    'l':    'l',    # ㄌ
    'h':    'h',    # ㄏ
    's':    's',    # ㄙ
    'si':   'si',   # ㄒ （注意放在 's' 前面）
}

# 只處理這兩種韻母，其他一律保留不動
TL_FINAL_MAP = {
    'onn': 'oonn',  # ㆧ
    'o':   'or',    # ㄜ
}

def convert_tl_to_bopo2(code: str) -> str:
    m = re.match(r'^([a-z]+)(\d+)$', code)
    if not m:
        return code
    body, tone = m.group(1), m.group(2)

    # 1. 聲母比對（從長到短）
    onset = ''
    rest  = body
    for key in sorted(TL_INITIAL_MAP, key=lambda x: -len(x)):
        if body.startswith(key):
            onset = TL_INITIAL_MAP[key]
            rest  = body[len(key):]
            break

    # 2. 韻母只對 TL_FINAL_MAP 裡的兩個 key 做轉換
    if rest in TL_FINAL_MAP:
        rest = TL_FINAL_MAP[rest]
    # 其餘 rest = 'ong','ok','om',... 都不會改變

    return f"{onset}{rest}{tone}"


def main(input_xlsx: str, output_xlsx: str = None):
    wb = load_workbook(input_xlsx)
    if '漢字庫' not in wb.sheetnames:
        print("錯誤：找不到工作表「漢字庫」")
        sys.exit(1)
    ws = wb['漢字庫']

    # 讀第一列標題，找出「台羅音標」和「注音二式」所在欄
    header = next(ws.iter_rows(min_row=1, max_row=1))
    cols = {cell.value: cell.col_idx for cell in header}
    if '台羅音標' not in cols:
        print("錯誤：找不到欄「台羅音標」")
        sys.exit(1)
    src_col = cols['台羅音標']

    if '注音二式' in cols:
        dst_col = cols['注音二式']
    else:
        dst_col = ws.max_column + 1
        ws.cell(row=1, column=dst_col, value='注音二式')

    # 從第 2 列開始，逐列轉換
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=src_col).value
        if raw is None:
            continue
        new = convert_tl_to_bopo2(str(raw).strip())
        ws.cell(row=row, column=dst_col, value=new)

    # 存檔
    save_path = output_xlsx or input_xlsx
    wb.save(save_path)
    print(f"已將結果寫入：{save_path}")

if __name__ == '__main__':
    if len(sys.argv) == 1:
        print("用法：python convert_tl_to_bopo2_excel.py input.xlsx [output.xlsx]")
        sys.exit(0)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) >= 3 else None
    main(inp, out)
